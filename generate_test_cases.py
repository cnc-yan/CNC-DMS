#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CNC-DMS 寮管理システム テストケース自動生成スクリプト
"""

import random
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 定数
# ============================================================
PROJECT_NAME = "CNC-DMS"
RANDOM_SUFFIX = str(random.randint(10000, 99999))
OUTPUT_DIR = r"c:\Users\Administrator\IdeaProjects\CNC-DMS\.trae\test"
FRONTEND_URL = "http://localhost:5173"
BACKEND_URL = "http://localhost:8080"
DB_INFO = "MySQL 8 (tbl_user, tbl_department, tbl_employee, tbl_dormitory, tbl_room, tbl_resident_record, tbl_billing, tbl_operation_log, tbl_change_history, tbl_import_log)"
PRD_DOC_NAME = "寮管理システム基本設計仕様書"

# ============================================================
# テストケースデータ
# ============================================================
# 各テストケース: (所属模块, 相关研发需求, 用例名称, 前置条件, 用例类型, 测试步骤, 预期结果, 优先级)

test_cases = []

# =====================================================================
# 1. 認証 (認証)
# =====================================================================
AUTH_PRECOND = f"""1. 前后端服务已启动（前端 {FRONTEND_URL}，后端 {BACKEND_URL}）
2. 数据库 {DB_INFO} 连接正常
3. 已存在管理员账号 admin/admin123"""

auth_module = "認証"
auth_req = "2. システム全体アーキテクチャ / 8. APIアーキテクチャ設計 / 16. セキュリティ設計"

test_cases.append((auth_module, auth_req,
    "正常登录_验证成功跳转首页",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [断言URL] /home\n7. [断言] 欢迎 可见\n8. [DB] SELECT userid,userName,enabled FROM tbl_user WHERE userid='admin'\n9. [截图] 登录成功-首页仪表盘",
    "1. 登录页面加载完成，显示用户名和密码输入框\n2. 用户名输入框显示admin\n3. 密码输入框显示密文\n4. 触发登录请求\n5. 页面跳转至首页\n6. URL包含/home\n7. 页面显示欢迎信息\n8. 数据库返回admin用户记录，enabled=true\n9. 截图保存成功",
    "P1"))

test_cases.append((auth_module, auth_req,
    "异常登录_验证错误密码提示",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::wrongpass\n4. [点击] 登录\n5. [等待] 1000\n6. [断言] 密码错误 可见 或 用户不存在 可见",
    "1. 登录页面加载完成\n2. 用户名输入框显示admin\n3. 密码输入框显示密文\n4. 触发登录请求\n5. 页面不跳转\n6. 显示错误提示信息（密码错误）",
    "P1"))

test_cases.append((auth_module, auth_req,
    "异常登录_验证不存在的用户拒绝访问",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::nonexist_user\n3. [输入] input[placeholder*=\"密码\"]::pass123\n4. [点击] 登录\n5. [等待] 1000\n6. [断言] 用户不存在 可见",
    "1. 登录页面加载完成\n2. 用户名输入框显示nonexist_user\n3. 密码输入框显示密文\n4. 触发登录请求\n5. 页面不跳转\n6. 显示用户不存在错误提示",
    "P1"))

test_cases.append((auth_module, auth_req,
    "边界登录_验证空账号密码提交",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [点击] 登录\n3. [等待] 1000\n4. [断言] 请输入 可见",
    "1. 登录页面加载完成\n2. 未填写内容直接点击登录\n3. 表单校验触发\n4. 页面提示请输入用户名/密码",
    "P2"))

test_cases.append((auth_module, auth_req,
    "异常登录_验证禁用用户拒绝登录",
    AUTH_PRECOND.replace("已存在管理员账号 admin/admin123", "已存在禁用用户 disabled_user/test123\n4. 执行SQL: UPDATE tbl_user SET enabled=0 WHERE userid='disabled_user'"),
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::disabled_user\n3. [输入] input[placeholder*=\"密码\"]::test123\n4. [点击] 登录\n5. [等待] 1000\n6. [断言] 禁用 可见 或 不可登录 可见",
    "1. 登录页面加载完成\n2. 用户名输入框显示disabled_user\n3. 密码输入框显示密文\n4. 触发登录请求\n5. 页面不跳转\n6. 显示账户已被禁用的提示",
    "P2"))

test_cases.append((auth_module, auth_req,
    "正常获取当前用户_验证已登录用户信息",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /home\n7. [断言] admin 可见 或 管理者 可见\n8. [截图] 登录后用户信息",
    "1. 登录页面加载完成\n2-3. 输入凭据\n4. 触发登录请求\n5. 页面跳转\n6. 导航到首页\n7. 页面显示当前登录用户信息（admin/管理者）\n8. 截图保存成功",
    "P2"))

# =====================================================================
# 2. ユーザー管理 (ユーザー管理)
# =====================================================================
user_module = "ユーザー管理"
user_req = "2.3/2.5 ユーザー管理CRUD API"

test_cases.append((user_module, user_req,
    "正常查询用户列表_验证默认加载",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /system/user\n7. [等待] 1000\n8. [断言] 用户管理 可见\n9. [DB] SELECT COUNT(*) as cnt FROM tbl_user",
    "1-4. 登录成功\n5. 页面跳转至首页\n6. 导航到用户管理页面\n7. 页面加载完成\n8. 页面标题显示用户管理\n9. 数据库返回用户总数",
    "P2"))

test_cases.append((user_module, user_req,
    "正常创建用户_验证新增用户成功",
    AUTH_PRECOND + "\n4. 登录用户具有用户创建权限",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /system/user\n7. [等待] 1000\n8. [点击] 新增\n9. [等待] 500\n10. [输入] .el-dialog input[placeholder*=\"用户\"]::testuser01\n11. [输入] .el-dialog input[placeholder*=\"密码\"]::Pass1234\n12. [输入] .el-dialog input[placeholder*=\"姓名\"]::测试用户\n13. [点击] 确定\n14. [等待] 1000\n15. [断言] testuser01 可见\n16. [DB] SELECT userid,userName,enabled FROM tbl_user WHERE userid='testuser01'",
    "1-5. 登录成功\n6-7. 导航到用户管理页面\n8. 新增对话框打开\n9. 对话框加载完成\n10-12. 填写表单\n13. 提交创建\n14. 处理请求\n15. 列表中显示testuser01\n16. 数据库返回新用户记录",
    "P2"))

test_cases.append((user_module, user_req,
    "异常创建用户_验证重复用户ID提示",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /system/user\n7. [等待] 1000\n8. [点击] 新增\n9. [等待] 500\n10. [输入] .el-dialog input[placeholder*=\"用户\"]::admin\n11. [输入] .el-dialog input[placeholder*=\"密码\"]::Pass1234\n12. [输入] .el-dialog input[placeholder*=\"姓名\"]::重复用户\n13. [点击] 确定\n14. [等待] 1000\n15. [断言] 已存在 可见 或 重复 可见",
    "1-5. 登录成功\n6-7. 导航到用户管理页面\n8. 新增对话框打开\n9-12. 填写已存在的用户ID\n13. 提交创建\n14. 处理请求\n15. 提示用户ID已存在",
    "P2"))

test_cases.append((user_module, user_req,
    "正常更新用户_验证更新后信息正确",
    AUTH_PRECOND + "\n4. 已存在用户 testuser01",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /system/user\n7. [等待] 1000\n8. [点击] testuser01 行的编辑按钮\n9. [等待] 500\n10. [输入] .el-dialog input[placeholder*=\"姓名\"]::更新后的姓名\n11. [点击] 确定\n12. [等待] 1000\n13. [断言] 更新后的姓名 可见\n14. [DB] SELECT userName FROM tbl_user WHERE userid='testuser01'",
    "1-5. 登录成功\n6-7. 导航到用户管理页面\n8. 编辑对话框打开\n9-10. 修改姓名\n11. 提交更新\n12. 处理请求\n13. 列表显示更新后的姓名\n14. 数据库返回更新后的姓名",
    "P2"))

test_cases.append((user_module, user_req,
    "正常删除用户_验证单用户删除",
    AUTH_PRECOND + "\n4. 已存在测试用户 testuser02",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /system/user\n7. [等待] 1000\n8. [点击] testuser02 行的删除按钮\n9. [等待] 500\n10. [点击] 确认\n11. [等待] 1000\n12. [断言] testuser02 不可见\n13. [DB] SELECT COUNT(*) as cnt FROM tbl_user WHERE userid='testuser02'",
    "1-5. 登录成功\n6-7. 导航到用户管理页面\n8. 删除确认对话框打开\n9-10. 确认删除\n11. 处理请求\n12. 列表中testuser02消失\n13. 数据库返回0条记录",
    "P2"))

test_cases.append((user_module, user_req,
    "正常一括削除用户_验证批量删除成功",
    AUTH_PRECOND + "\n4. 已存在测试用户 testdel1, testdel2",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /system/user\n7. [等待] 1000\n8. [点击] testdel1 复选框\n9. [点击] testdel2 复选框\n10. [点击] 批量删除\n11. [等待] 500\n12. [点击] 确认\n13. [等待] 1000\n14. [DB] SELECT COUNT(*) as cnt FROM tbl_user WHERE userid IN ('testdel1','testdel2')",
    "1-5. 登录成功\n6-7. 导航到用户管理页面\n8-9. 勾选两个用户\n10. 点击批量删除\n11-12. 确认删除\n13. 处理请求\n14. 数据库返回0条记录，两个用户均被删除",
    "P2"))

test_cases.append((user_module, user_req,
    "边界查询用户_验证空搜索条件",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /system/user\n7. [等待] 1000\n8. [输入] input[placeholder*=\"用户ID\"]::\n9. [输入] input[placeholder*=\"姓名\"]::\n10. [点击] 搜索\n11. [等待] 1000\n12. [断言] 共 可见",
    "1-5. 登录成功\n6-7. 导航到用户管理页面\n8-9. 搜索条件为空\n10. 点击搜索\n11. 页面刷新\n12. 显示所有用户列表，包含总数信息",
    "P3"))

# =====================================================================
# 3. 部門管理 (部門管理)
# =====================================================================
dept_module = "部門管理"
dept_req = "3. 部門管理 API"

test_cases.append((dept_module, dept_req,
    "正常查询部门列表_验证默认加载",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/department\n7. [等待] 1000\n8. [断言] 部门管理 可见\n9. [DB] SELECT COUNT(*) as cnt FROM tbl_department WHERE status=1",
    "1-4. 登录成功\n5. 页面跳转至首页\n6. 导航到部门管理页面\n7. 页面加载完成\n8. 页面标题显示部门管理\n9. 数据库返回有效部门数量",
    "P2"))

test_cases.append((dept_module, dept_req,
    "正常创建部门_验证新增部门成功",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/department\n7. [等待] 1000\n8. [点击] 新增\n9. [等待] 500\n10. [输入] .el-dialog input[placeholder*=\"部门代码\"]::DEPT-099\n11. [输入] .el-dialog input[placeholder*=\"部门名称\"]::测试部门\n12. [选择] .el-dialog .el-select::东京\n13. [点击] 确定\n14. [等待] 1000\n15. [断言] DEPT-099 可见\n16. [DB] SELECT dept_code,dept_name,region FROM tbl_department WHERE dept_code='DEPT-099'",
    "1-5. 登录成功\n6-7. 导航到部门管理页面\n8. 新增对话框打开\n9-12. 填写部门信息\n13. 提交创建\n14. 处理请求\n15. 列表显示新部门\n16. 数据库返回新部门记录",
    "P2"))

test_cases.append((dept_module, dept_req,
    "异常创建部门_验证重复部门代码提示",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/department\n7. [等待] 1000\n8. [点击] 新增\n9. [等待] 500\n10. [输入] .el-dialog input[placeholder*=\"部门代码\"]::DEPT-001\n11. [输入] .el-dialog input[placeholder*=\"部门名称\"]::重复部门\n12. [点击] 确定\n13. [等待] 1000\n14. [断言] 已存在 可见 或 重复 可见",
    "1-5. 登录成功\n6-7. 导航到部门管理页面\n8. 新增对话框打开\n9-11. 填写已存在的部门代码\n12. 提交创建\n13. 处理请求\n14. 提示部门代码已存在",
    "P2"))

test_cases.append((dept_module, dept_req,
    "正常更新部门_验证更新后信息正确",
    AUTH_PRECOND + "\n4. 已存在部门 DEPT-099",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/department\n7. [等待] 1000\n8. [点击] DEPT-099 行的编辑按钮\n9. [等待] 500\n10. [输入] .el-dialog input[placeholder*=\"部门名称\"]::更新后的部门名\n11. [点击] 确定\n12. [等待] 1000\n13. [断言] 更新后的部门名 可见\n14. [DB] SELECT dept_name FROM tbl_department WHERE dept_code='DEPT-099'",
    "1-5. 登录成功\n6-7. 导航到部门管理页面\n8. 编辑对话框打开\n9-10. 修改部门名称\n11. 提交更新\n12. 处理请求\n13. 列表显示更新后的名称\n14. 数据库返回更新后的部门名称",
    "P2"))

test_cases.append((dept_module, dept_req,
    "正常删除部门_验证删除成功",
    AUTH_PRECOND + "\n4. 已存在测试部门 DEPT-099",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/department\n7. [等待] 1000\n8. [点击] DEPT-099 行的删除按钮\n9. [等待] 500\n10. [点击] 确认\n11. [等待] 1000\n12. [断言] DEPT-099 不可见\n13. [DB] SELECT status FROM tbl_department WHERE dept_code='DEPT-099'",
    "1-5. 登录成功\n6-7. 导航到部门管理页面\n8. 删除确认对话框打开\n9-10. 确认删除\n11. 处理请求\n12. 列表中DEPT-099消失\n13. 数据库中状态标记为无效(status=0)",
    "P2"))

test_cases.append((dept_module, dept_req,
    "正常查询部门_验证条件搜索功能",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/department\n7. [等待] 1000\n8. [输入] input[placeholder*=\"地域\"]::东京\n9. [点击] 搜索\n10. [等待] 1000\n11. [断言] 东京本社 可见\n12. [DB] SELECT dept_code,dept_name FROM tbl_department WHERE region='东京' AND status=1",
    "1-5. 登录成功\n6-7. 导航到部门管理页面\n8-9. 输入地域搜索条件\n10. 页面刷新\n11. 列表显示东京地区的部门\n12. 数据库返回东京地区有效部门列表",
    "P3"))

# =====================================================================
# 4. 社員管理 (社員管理)
# =====================================================================
emp_module = "社員管理"
emp_req = "4. 社員管理 API"

test_cases.append((emp_module, emp_req,
    "正常查询员工列表_验证默认加载分页",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/employee\n7. [等待] 1000\n8. [断言] 员工管理 可见\n9. [DB] SELECT COUNT(*) as cnt FROM tbl_employee WHERE emp_status=1",
    "1-5. 登录成功\n6-7. 导航到员工管理页面\n8. 页面标题显示员工管理\n9. 数据库返回在职员工总数",
    "P2"))

test_cases.append((emp_module, emp_req,
    "正常创建员工_验证新增员工成功",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/employee\n7. [等待] 1000\n8. [点击] 新增\n9. [等待] 500\n10. [输入] .el-dialog input[placeholder*=\"工号\"]::EMP-99999\n11. [输入] .el-dialog input[placeholder*=\"姓名\"]::测试员工\n12. [选择] .el-dialog .el-select[placeholder*=\"性别\"]::MALE\n13. [输入] .el-dialog input[placeholder*=\"国籍\"]::日本\n14. [输入] .el-dialog input[placeholder*=\"入社日\"]::2026-01-01\n15. [点击] 确定\n16. [等待] 1000\n17. [断言] EMP-99999 可见\n18. [DB] SELECT emp_no,emp_name,gender FROM tbl_employee WHERE emp_no='EMP-99999'",
    "1-5. 登录成功\n6-7. 导航到员工管理页面\n8. 新增对话框打开\n9-14. 填写员工信息\n15. 提交创建\n16. 处理请求\n17. 列表显示新员工\n18. 数据库返回新员工记录",
    "P2"))

test_cases.append((emp_module, emp_req,
    "异常创建员工_验证重复工号提示",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/employee\n7. [等待] 1000\n8. [点击] 新增\n9. [等待] 500\n10. [输入] .el-dialog input[placeholder*=\"工号\"]::EMP-00001\n11. [输入] .el-dialog input[placeholder*=\"姓名\"]::重复员工\n12. [点击] 确定\n13. [等待] 1000\n14. [断言] 已存在 可见 或 重复 可见",
    "1-5. 登录成功\n6-7. 导航到员工管理页面\n8. 新增对话框打开\n9-11. 填写已存在的工号\n12. 提交创建\n13. 处理请求\n14. 提示工号已存在",
    "P2"))

test_cases.append((emp_module, emp_req,
    "正常更新员工_验证更新后信息正确",
    AUTH_PRECOND + "\n4. 已存在员工 EMP-99999",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/employee\n7. [等待] 1000\n8. [点击] EMP-99999 行的编辑按钮\n9. [等待] 500\n10. [输入] .el-dialog input[placeholder*=\"姓名\"]::更新后的员工名\n11. [点击] 确定\n12. [等待] 1000\n13. [断言] 更新后的员工名 可见\n14. [DB] SELECT emp_name FROM tbl_employee WHERE emp_no='EMP-99999'",
    "1-5. 登录成功\n6-7. 导航到员工管理页面\n8. 编辑对话框打开\n9-10. 修改员工姓名\n11. 提交更新\n12. 处理请求\n13. 列表显示更新后的姓名\n14. 数据库返回更新后的员工姓名",
    "P2"))

test_cases.append((emp_module, emp_req,
    "正常删除员工_验证删除成功",
    AUTH_PRECOND + "\n4. 已存在测试员工 EMP-99999",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/employee\n7. [等待] 1000\n8. [点击] EMP-99999 行的删除按钮\n9. [等待] 500\n10. [点击] 确认\n11. [等待] 1000\n12. [断言] EMP-99999 不可见\n13. [DB] SELECT emp_status FROM tbl_employee WHERE emp_no='EMP-99999'",
    "1-5. 登录成功\n6-7. 导航到员工管理页面\n8. 删除确认对话框打开\n9-10. 确认删除\n11. 处理请求\n12. 列表中EMP-99999消失\n13. 数据库将员工状态标记为离职(emp_status=0)",
    "P2"))

test_cases.append((emp_module, emp_req,
    "边界查询员工_验证多种条件组合搜索",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/employee\n7. [等待] 1000\n8. [输入] input[placeholder*=\"姓名\"]::山田\n9. [选择] .el-select[placeholder*=\"状态\"]::在職\n10. [点击] 搜索\n11. [等待] 1000\n12. [断言] 山田 可见\n13. [DB] SELECT emp_no,emp_name FROM tbl_employee WHERE emp_name LIKE '%山田%' AND emp_status=1",
    "1-5. 登录成功\n6-7. 导航到员工管理页面\n8-9. 输入姓名和状态搜索条件\n10. 点击搜索\n11. 页面刷新\n12. 列表显示符合条件的员工\n13. 数据库返回姓名含山田且状态为在職的员工",
    "P3"))

# =====================================================================
# 5. 寮管理 (寮管理)
# =====================================================================
dorm_module = "寮管理"
dorm_req = "4.1 寮管理 / 5. 寮管理API"

test_cases.append((dorm_module, dorm_req,
    "正常查询寮列表_验证默认加载",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/dormitory\n7. [等待] 1000\n8. [断言] 寮管理 可见\n9. [DB] SELECT COUNT(*) as cnt FROM tbl_dormitory WHERE status=1",
    "1-5. 登录成功\n6-7. 导航到寮管理页面\n8. 页面标题显示寮管理\n9. 数据库返回运营中的寮数量",
    "P2"))

test_cases.append((dorm_module, dorm_req,
    "正常创建寮_验证新增寮成功",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/dormitory\n7. [等待] 1000\n8. [点击] 新增\n9. [等待] 500\n10. [输入] .el-dialog input[placeholder*=\"寮名称\"]::测试寮\n11. [输入] .el-dialog input[placeholder*=\"地域\"]::东京\n12. [选择] .el-dialog .el-select[placeholder*=\"入居条件\"]::ANY\n13. [输入] .el-dialog input[placeholder*=\"住所\"]::东京测试地址\n14. [点击] 确定\n15. [等待] 1000\n16. [断言] 测试寮 可见\n17. [DB] SELECT dorm_name,region,dorm_condition FROM tbl_dormitory WHERE dorm_name='测试寮'",
    "1-5. 登录成功\n6-7. 导航到寮管理页面\n8. 新增对话框打开\n9-13. 填写寮信息\n14. 提交创建\n15. 处理请求\n16. 列表显示新寮\n17. 数据库返回新寮记录",
    "P2"))

test_cases.append((dorm_module, dorm_req,
    "异常创建寮_验证必填字段校验",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/dormitory\n7. [等待] 1000\n8. [点击] 新增\n9. [等待] 500\n10. [点击] 确定\n11. [等待] 1000\n12. [断言] 必填 可见 或 请输入 可见",
    "1-5. 登录成功\n6-7. 导航到寮管理页面\n8. 新增对话框打开\n9-10. 不填任何信息直接提交\n11. 表单校验触发\n12. 提示寮名称为必填项",
    "P3"))

test_cases.append((dorm_module, dorm_req,
    "正常更新寮_验证寮信息更新成功",
    AUTH_PRECOND + "\n4. 已存在测试寮 测试寮",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/dormitory\n7. [等待] 1000\n8. [点击] 测试寮 行的编辑按钮\n9. [等待] 500\n10. [输入] .el-dialog input[placeholder*=\"寮名称\"]::更新后的测试寮\n11. [点击] 确定\n12. [等待] 1000\n13. [断言] 更新后的测试寮 可见\n14. [DB] SELECT dorm_name FROM tbl_dormitory WHERE id=LAST_INSERT_ID()",
    "1-5. 登录成功\n6-7. 导航到寮管理页面\n8. 编辑对话框打开\n9-10. 修改寮名称\n11. 提交更新\n12. 处理请求\n13. 列表显示更新后的名称\n14. 数据库返回更新后的寮名称",
    "P2"))

test_cases.append((dorm_module, dorm_req,
    "正常删除寮_验证删除成功",
    AUTH_PRECOND + "\n4. 已存在测试寮（无关联房间）",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/dormitory\n7. [等待] 1000\n8. [点击] 要删除的寮行删除按钮\n9. [等待] 500\n10. [点击] 确认\n11. [等待] 1000\n12. [断言] 删除成功 可见\n13. [DB] SELECT status FROM tbl_dormitory WHERE id={id}",
    "1-5. 登录成功\n6-7. 导航到寮管理页面\n8. 删除确认对话框打开\n9-10. 确认删除\n11. 处理请求\n12. 页面提示删除成功\n13. 数据库将寮状态标记为停止中(status=0)",
    "P2"))

test_cases.append((dorm_module, dorm_req,
    "正常查询寮_验证按地域条件搜索",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/dormitory\n7. [等待] 1000\n8. [选择] .el-select[placeholder*=\"地域\"]::东京\n9. [点击] 搜索\n10. [等待] 1000\n11. [断言] 东京 可见\n12. [DB] SELECT dorm_name FROM tbl_dormitory WHERE region='东京' AND status=1",
    "1-5. 登录成功\n6-7. 导航到寮管理页面\n8-9. 选择地域条件搜索\n10. 页面刷新\n11. 列表显示东京地区的寮\n12. 数据库返回东京地区的运营中寮",
    "P3"))

# =====================================================================
# 6. 部屋管理 (部屋管理)
# =====================================================================
room_module = "部屋管理"
room_req = "4.2 部屋管理 / 6. 部屋管理API"

test_cases.append((room_module, room_req,
    "正常查询房间列表_验证默认加载",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/room\n7. [等待] 1000\n8. [断言] 房间管理 可见\n9. [DB] SELECT COUNT(*) as cnt FROM tbl_room WHERE status IN (1,2)",
    "1-5. 登录成功\n6-7. 导航到房间管理页面\n8. 页面标题显示房间管理\n9. 数据库返回使用中的房间总数",
    "P2"))

test_cases.append((room_module, room_req,
    "正常创建房间_验证新增房间成功",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/room\n7. [等待] 1000\n8. [点击] 新增\n9. [等待] 500\n10. [选择] .el-dialog .el-select[placeholder*=\"所属寮\"]::東京第一寮\n11. [输入] .el-dialog input[placeholder*=\"房间号\"]::999\n12. [输入] .el-dialog input[placeholder*=\"定员\"]::1\n13. [输入] .el-dialog input[placeholder*=\"日额\"]::1500\n14. [点击] 确定\n15. [等待] 1000\n16. [断言] 999 可见\n17. [DB] SELECT r.room_number,r.capacity,r.daily_rate FROM tbl_room r JOIN tbl_dormitory d ON r.dormitory_id=d.id WHERE r.room_number='999' AND d.dorm_name='東京第一寮'",
    "1-5. 登录成功\n6-7. 导航到房间管理页面\n8. 新增对话框打开\n9-13. 填写房间信息\n14. 提交创建\n15. 处理请求\n16. 列表显示新房间\n17. 数据库返回新房间记录",
    "P2"))

test_cases.append((room_module, room_req,
    "异常创建房间_验证同一寮下重复房间号提示",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/room\n7. [等待] 1000\n8. [点击] 新增\n9. [等待] 500\n10. [选择] .el-dialog .el-select[placeholder*=\"所属寮\"]::東京第一寮\n11. [输入] .el-dialog input[placeholder*=\"房间号\"]::101\n12. [点击] 确定\n13. [等待] 1000\n14. [断言] 已存在 可见 或 重复 可见",
    "1-5. 登录成功\n6-7. 导航到房间管理页面\n8. 新增对话框打开\n9-11. 填写已存在的房间号\n12. 提交创建\n13. 处理请求\n14. 提示房间号已存在",
    "P2"))

test_cases.append((room_module, room_req,
    "正常更新房间_验证更新房间信息成功",
    AUTH_PRECOND + "\n4. 已存在房间号999（测试用）",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/room\n7. [等待] 1000\n8. [点击] 房间999行的编辑按钮\n9. [等待] 500\n10. [输入] .el-dialog input[placeholder*=\"日额\"]::2000\n11. [点击] 确定\n12. [等待] 1000\n13. [断言] 2,000 可见 或 2000 可见\n14. [DB] SELECT daily_rate FROM tbl_room WHERE room_number='999'",
    "1-5. 登录成功\n6-7. 导航到房间管理页面\n8. 编辑对话框打开\n9-10. 修改日额\n11. 提交更新\n12. 处理请求\n13. 列表显示更新后的金额\n14. 数据库返回更新后的日额",
    "P2"))

test_cases.append((room_module, room_req,
    "正常删除房间_验证删除成功",
    AUTH_PRECOND + "\n4. 已存在测试房间（无关联入住记录）",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/room\n7. [等待] 1000\n8. [点击] 要删除的房间行删除按钮\n9. [等待] 500\n10. [点击] 确认\n11. [等待] 1000\n12. [断言] 删除成功 可见\n13. [DB] SELECT status FROM tbl_room WHERE id={id}",
    "1-5. 登录成功\n6-7. 导航到房间管理页面\n8. 删除确认对话框打开\n9-10. 确认删除\n11. 处理请求\n12. 页面提示删除成功\n13. 数据库将房间状态标记为使用不可(status=0)",
    "P2"))

test_cases.append((room_module, room_req,
    "正常查询房间_验证按寮ID过滤",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/room\n7. [等待] 1000\n8. [选择] .el-select[placeholder*=\"所属寮\"]::東京第一寮\n9. [点击] 搜索\n10. [等待] 1000\n11. [断言] 101 可见\n12. [断言] 大阪 不可见\n13. [DB] SELECT r.room_number FROM tbl_room r JOIN tbl_dormitory d ON r.dormitory_id=d.id WHERE d.dorm_name='東京第一寮'",
    "1-5. 登录成功\n6-7. 导航到房间管理页面\n8-9. 选择所属寮过滤\n10. 页面刷新\n11. 列表显示东京第一寮的房间\n12. 不显示其他寮的房间\n13. 数据库返回东京第一寮的所有房间",
    "P3"))

# =====================================================================
# 7. 入居履歴管理 (入居管理)
# =====================================================================
resident_module = "入居履歴管理"
resident_req = "5. 入退寮管理 / 7. 入居履歴API / 10.2 インポートバリデーションルール"

test_cases.append((resident_module, resident_req,
    "正常查询入居履历列表_验证默认加载",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /resident/records\n7. [等待] 1000\n8. [断言] 入居履历 可见\n9. [DB] SELECT COUNT(*) as cnt FROM tbl_resident_record WHERE is_active=1",
    "1-5. 登录成功\n6-7. 导航到入居履历管理页面\n8. 页面标题显示入居履历\n9. 数据库返回当前入居中的记录数",
    "P1"))

test_cases.append((resident_module, resident_req,
    "正常入居登録_验证新增入居成功",
    AUTH_PRECOND + "\n4. 存在空房间（如东京第一寮201号室, room_id=4）\n5. 存在未入居的员工",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /resident/records\n7. [等待] 1000\n8. [点击] 入居登録\n9. [等待] 500\n10. [选择] .el-dialog .el-select[placeholder*=\"员工\"]::佐藤次郎\n11. [选择] .el-dialog .el-select[placeholder*=\"房间\"]::203\n12. [输入] .el-dialog input[placeholder*=\"入居日\"]::2026-06-20\n13. [点击] 确定\n14. [等待] 1000\n15. [断言] 入居成功 可见\n16. [DB] SELECT e.emp_name,r.room_number,rr.checkin_date,rr.is_active FROM tbl_resident_record rr JOIN tbl_employee e ON rr.employee_id=e.id JOIN tbl_room r ON rr.room_id=r.id WHERE e.emp_no='EMP-00003' AND rr.is_active=1",
    "1-5. 登录成功\n6-7. 导航到入居履历页面\n8. 入居登録对话框打开\n9-12. 选择员工、房间和入居日\n13. 提交入居\n14. 处理请求\n15. 页面提示入居成功\n16. 数据库返回入居记录，is_active=1",
    "P1"))

test_cases.append((resident_module, resident_req,
    "异常入居登録_验证同一员工重复入居提示",
    AUTH_PRECOND + "\n4. 员工山田太郎(emp_id=1)已入居中",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /resident/records\n7. [等待] 1000\n8. [点击] 入居登録\n9. [等待] 500\n10. [选择] .el-dialog .el-select[placeholder*=\"员工\"]::山田太郎\n11. [选择] .el-dialog .el-select[placeholder*=\"房间\"]::203\n12. [输入] .el-dialog input[placeholder*=\"入居日\"]::2026-06-20\n13. [点击] 确定\n14. [等待] 1000\n15. [断言] 入居中 可见 或 重复 可见 或 既に 可见",
    "1-5. 登录成功\n6-7. 导航到入居履历页面\n8. 入居登録对话框打开\n9-12. 选择已入居的员工\n13. 提交入居\n14. 处理请求\n15. 提示该员工已入居中",
    "P1"))

test_cases.append((resident_module, resident_req,
    "异常入居登録_验证性别不匹配拒绝入居",
    AUTH_PRECOND + "\n4. 女性员工已存在\n5. 东京第一寮是男性寮(MALE)",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /resident/records\n7. [等待] 1000\n8. [点击] 入居登録\n9. [等待] 500\n10. [选择] .el-dialog .el-select[placeholder*=\"员工\"]::田中花子\n11. [选择] .el-dialog .el-select[placeholder*=\"房间\"]::101\n12. [输入] .el-dialog input[placeholder*=\"入居日\"]::2026-06-20\n13. [点击] 确定\n14. [等待] 1000\n15. [断言] 性别 可见 或 不一致 可见",
    "1-5. 登录成功\n6-7. 导航到入居履历页面\n8. 入居登録对话框打开\n9-12. 选择女性员工和男性寮的房间\n13. 提交入居\n14. 处理请求\n15. 提示性别与寮条件不匹配",
    "P1"))

test_cases.append((resident_module, resident_req,
    "正常退寮処理_验证退寮成功及费用自动计算",
    AUTH_PRECOND + "\n4. 存在入居中的记录（如佐藤次郎入居）",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /resident/records\n7. [等待] 1000\n8. [点击] 入居中的员工行退寮按钮\n9. [等待] 500\n10. [输入] .el-dialog input[placeholder*=\"退寮日\"]::2026-06-30\n11. [输入] .el-dialog textarea[placeholder*=\"理由\"]::退寮测试\n12. [点击] 确定\n13. [等待] 1000\n14. [断言] 退寮成功 可见\n15. [DB] SELECT checkout_date,is_active,total_fee FROM tbl_resident_record WHERE id={id}",
    "1-5. 登录成功\n6-7. 导航到入居履历页面\n8. 退寮对话框打开\n9-11. 填写退寮日和理由\n12. 提交退寮\n13. 处理请求\n14. 页面提示退寮成功\n15. 数据库返回退寮日、is_active=0、total_fee自动计算",
    "P1"))

test_cases.append((resident_module, resident_req,
    "异常退寮処理_验证已退寮记录不可重复退寮",
    AUTH_PRECOND + "\n4. 存在已退寮的记录",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /resident/records\n7. [等待] 1000\n8. [选择] .el-select[placeholder*=\"状态\"]::退寮済\n9. [点击] 搜索\n10. [等待] 1000\n11. [点击] 退寮済记录行退寮按钮\n12. [等待] 1000\n13. [断言] 退寮済 可见 或 不可 可见",
    "1-5. 登录成功\n6-7. 导航到入居履历页面\n8-9. 搜索退寮済记录\n10. 页面刷新\n11-12. 尝试对已退寮记录操作\n13. 系统不允许重复退寮",
    "P2"))

test_cases.append((resident_module, resident_req,
    "正常入居履歴削除_验证删除成功",
    AUTH_PRECOND + "\n4. 存在测试入居记录（可删除）",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /resident/records\n7. [等待] 1000\n8. [点击] 目标记录行删除按钮\n9. [等待] 500\n10. [点击] 确认\n11. [等待] 1000\n12. [断言] 删除成功 可见\n13. [DB] SELECT COUNT(*) as cnt FROM tbl_resident_record WHERE id={id}",
    "1-5. 登录成功\n6-7. 导航到入居履历页面\n8. 删除确认对话框打开\n9-10. 确认删除\n11. 处理请求\n12. 页面提示删除成功\n13. 数据库返回0条（物理删除或逻辑删除）",
    "P3"))

# =====================================================================
# 8. 寮割カレンダー (寮割カレンダー)
# =====================================================================
calendar_module = "寮割カレンダー"
calendar_req = "11. 画面およびインタラクション設計 / 11.2 ログ監査画面 / 寮割カレンダー画面"

test_cases.append((calendar_module, calendar_req,
    "正常显示カレンダー_验证默认当前月加载",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /resident/calendar\n7. [等待] 1000\n8. [断言] カレンダー 可见\n9. [截图] カレンダー默认加载",
    "1-5. 登录成功\n6-7. 导航到寮割カレンダー页面\n8. 页面标题显示カレンダー\n9. 截图保存成功",
    "P2"))

test_cases.append((calendar_module, calendar_req,
    "正常月切替_验证前月次月切换功能",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /resident/calendar\n7. [等待] 1000\n8. [点击] 前月\n9. [等待] 500\n10. [截图] 前月カレンダー\n11. [点击] 次月\n12. [等待] 500\n13. [截图] 次月カレンダー\n14. [点击] 今日\n15. [等待] 500\n16. [断言] 当月日期 可见",
    "1-5. 登录成功\n6-7. 导航到カレンダー页面\n8-9. 切换到前月\n10. 截图保存\n11-12. 切换到次月\n13. 截图保存\n14-15. 返回当月\n16. 显示当月日期信息",
    "P2"))

test_cases.append((calendar_module, calendar_req,
    "正常地域フィルタ_验证按地域筛选显示",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /resident/calendar\n7. [等待] 1000\n8. [选择] .el-select[placeholder*=\"地域\"]::东京\n9. [等待] 1000\n10. [断言] 东京 可见\n11. [截图] 东京地域カレンダー",
    "1-5. 登录成功\n6-7. 导航到カレンダー页面\n8-9. 选择地域过滤\n10. 页面显示东京地区的入居信息\n11. 截图保存成功",
    "P3"))

test_cases.append((calendar_module, calendar_req,
    "正常セル表示_验证入居者信息工具提示",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /resident/calendar\n7. [等待] 1000\n8. [悬停] .calendar-cell 第一个黄色单元格\n9. [等待] 500\n10. [断言] 入居者名 可见 或 山田 可见\n11. [截图] 工具提示内容",
    "1-5. 登录成功\n6-7. 导航到カレンダー页面\n8-9. 悬停在入居开始日单元格上\n10. 工具提示显示入居者详细信息\n11. 截图保存成功",
    "P2"))

test_cases.append((calendar_module, calendar_req,
    "正常退寮警告_验证14日内退寮红色标记",
    AUTH_PRECOND + "\n4. 存在14天内退寮的记录（如入居者退寮日接近当日）",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /resident/calendar\n7. [等待] 1000\n8. [断言] .el-date-table__row td.is-warning 可见 或 红色背景 可见\n9. [截图] 退寮警告显示",
    "1-5. 登录成功\n6-7. 导航到カレンダー页面\n8. 14日内退寮的单元格显示红色背景警告\n9. 截图保存成功",
    "P2"))

# =====================================================================
# 9. 寮費管理 (寮費管理)
# =====================================================================
billing_module = "寮費管理"
billing_req = "6. 寮費管理 / 5.3 寮費プロセス / 9. 利用料金確認API"

test_cases.append((billing_module, billing_req,
    "正常查询寮费列表_验证默认加载",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /billing\n7. [等待] 1000\n8. [断言] 寮费管理 可见\n9. [DB] SELECT COUNT(*) as cnt FROM tbl_billing",
    "1-5. 登录成功\n6-7. 导航到寮费管理页面\n8. 页面标题显示寮费管理\n9. 数据库返回寮费设定记录数",
    "P2"))

test_cases.append((billing_module, billing_req,
    "正常寮費確定_验证状态流转",
    AUTH_PRECOND + "\n4. 存在未确定状态的寮费记录",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /billing\n7. [等待] 1000\n8. [点击] 目标记录行确定按钮\n9. [等待] 500\n10. [点击] 确认\n11. [等待] 1000\n12. [断言] 确定成功 可见 或 已确定 可见\n13. [DB] SELECT status FROM tbl_billing WHERE id={id}",
    "1-5. 登录成功\n6-7. 导航到寮费管理页面\n8. 确定确认对话框打开\n9-10. 确认确定\n11. 处理请求\n12. 页面提示确定成功\n13. 数据库返回status更新为已确定状态",
    "P2"))

test_cases.append((billing_module, billing_req,
    "异常寮費確定_验证已锁定不可修改",
    AUTH_PRECOND + "\n4. 存在已锁定的寮费记录",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /billing\n7. [等待] 1000\n8. [点击] 已锁定记录行编辑按钮\n9. [等待] 1000\n10. [断言] 不可编辑 可见 或 已锁定 可见",
    "1-5. 登录成功\n6-7. 导航到寮费管理页面\n8. 尝试编辑已锁定记录\n9-10. 系统提示记录已锁定不可编辑",
    "P2"))

test_cases.append((billing_module, billing_req,
    "正常更新寮費設定_验证日额更新成功",
    AUTH_PRECOND + "\n4. 存在寮费设定记录（未锁定）",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /billing\n7. [等待] 1000\n8. [点击] 目标记录行编辑按钮\n9. [等待] 500\n10. [输入] .el-dialog input[placeholder*=\"日额\"]::2000\n11. [点击] 确定\n12. [等待] 1000\n13. [断言] 2,000 可见\n14. [DB] SELECT daily_rate FROM tbl_billing WHERE id={id}",
    "1-5. 登录成功\n6-7. 导航到寮费管理页面\n8. 编辑对话框打开\n9-10. 修改日额\n11. 提交更新\n12. 处理请求\n13. 列表显示更新后的金额\n14. 数据库返回更新后的日额",
    "P2"))

test_cases.append((billing_module, billing_req,
    "正常使用料金確認_验证费用计算正确",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /api/resident-records/usage-fees\n7. [等待] 1000\n8. [断言] totalFee 可见\n9. [DB] SELECT id,employee_id,room_id,total_fee FROM tbl_resident_record WHERE is_active=0 AND total_fee IS NOT NULL LIMIT 1",
    "1-5. 登录成功\n6-7. 访问利用料金確認API\n8. 返回数据包含totalFee字段\n9. 数据库返回已退寮记录的totalFee值",
    "P2"))

test_cases.append((billing_module, billing_req,
    "正常一括削除寮費_验证批量删除",
    AUTH_PRECOND + "\n4. 存在多条寮费设定记录（未锁定）",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /billing\n7. [等待] 1000\n8. [点击] 多条记录复选框\n9. [点击] 批量删除\n10. [等待] 500\n11. [点击] 确认\n12. [等待] 1000\n13. [断言] 删除成功 可见\n14. [DB] SELECT COUNT(*) as cnt FROM tbl_billing WHERE id IN (...)",
    "1-5. 登录成功\n6-7. 导航到寮费管理页面\n8-9. 勾选多条记录点击批量删除\n10-11. 确认删除\n12. 处理请求\n13. 页面提示删除成功\n14. 数据库记录已删除",
    "P3"))

# =====================================================================
# 10. 空き室管理 (空き室管理)
# =====================================================================
vacancy_module = "空き室管理"
vacancy_req = "8. 空き室管理 / 8.2 機能要件"

test_cases.append((vacancy_module, vacancy_req,
    "正常显示空室一覧_验证默认加载",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /vacancy\n7. [等待] 1000\n8. [断言] 空室 可见\n9. [DB] SELECT r.id,r.room_number,r.capacity,r.current_occupancy,d.dorm_name FROM tbl_room r JOIN tbl_dormitory d ON r.dormitory_id=d.id WHERE r.status=1 AND r.capacity > r.current_occupancy",
    "1-5. 登录成功\n6-7. 导航到空室管理页面\n8. 页面标题显示空室信息\n9. 数据库返回所有空室（有空余床位的房间）",
    "P2"))

test_cases.append((vacancy_module, vacancy_req,
    "正常空室検索_验证性别条件筛选",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /vacancy\n7. [等待] 1000\n8. [选择] .el-select[placeholder*=\"性别\"]::MALE\n9. [等待] 1000\n10. [断言] 男性 可见 或 MALE 可见\n11. [DB] SELECT d.dorm_name,d.dorm_condition FROM tbl_dormitory d WHERE d.dorm_condition='MALE' AND d.status=1",
    "1-5. 登录成功\n6-7. 导航到空室管理页面\n8-9. 选择男性条件筛选\n10. 页面显示男性寮的空室信息\n11. 数据库返回男性寮列表",
    "P2"))

test_cases.append((vacancy_module, vacancy_req,
    "正常空室検索_验证寮指定筛选",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /vacancy\n7. [等待] 1000\n8. [选择] .el-select[placeholder*=\"寮\"]::東京第一寮\n9. [等待] 1000\n10. [断言] 101 可见\n11. [断言] 大阪 不可见\n12. [DB] SELECT room_number,status FROM tbl_room WHERE dormitory_id=1 AND capacity > current_occupancy",
    "1-5. 登录成功\n6-7. 导航到空室管理页面\n8-9. 选择指定寮\n10. 页面显示该寮的空室\n11. 不显示其他寮的房间\n12. 数据库返回该寮的空室列表",
    "P3"))

test_cases.append((vacancy_module, vacancy_req,
    "正常空室容量表示_验证颜色区分",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /vacancy\n7. [等待] 1000\n8. [断言] 空室数 可见 或 容量 可见\n9. [截图] 空室容量颜色显示",
    "1-5. 登录成功\n6-7. 导航到空室管理页面\n8. 页面显示空室容量信息\n9. 截图保存（验证不同容量的颜色区分）",
    "P3"))

test_cases.append((vacancy_module, vacancy_req,
    "正常空室判定_验证入居后状态实时更新",
    AUTH_PRECOND + "\n4. 存在空房间（如room_id=4, capacity=2, current_occupancy=0）",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /vacancy\n7. [等待] 1000\n8. [截图] 入居前空室状态\n9. [导航] /resident/records\n10. [点击] 入居登録\n11. [选择] 员工和空房间\n12. [输入] 入居日\n13. [点击] 确定\n14. [等待] 1000\n15. [导航] /vacancy\n16. [等待] 1000\n17. [截图] 入居后空室状态\n18. [DB] SELECT current_occupancy FROM tbl_room WHERE id={room_id}",
    "1-5. 登录成功\n6-8. 记录入居前空室状态\n9-13. 执行入居操作\n14. 入居成功\n15-16. 再次查看空室\n17. 截图保存入居后状态\n18. 数据库返回更新后的入居人数",
    "P1"))

# =====================================================================
# 11. 操作ログ (操作ログ)
# =====================================================================
oplog_module = "操作ログ"
oplog_req = "15. ログおよび監査設計 / 11.2 ログ監査画面"

test_cases.append((oplog_module, oplog_req,
    "正常查询操作日志列表_验证默认加载",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /system/operation-log\n7. [等待] 1000\n8. [断言] 操作日志 可见\n9. [DB] SELECT COUNT(*) as cnt FROM tbl_operation_log",
    "1-5. 登录成功\n6-7. 导航到操作日志页面\n8. 页面标题显示操作日志\n9. 数据库返回操作日志总数",
    "P3"))

test_cases.append((oplog_module, oplog_req,
    "正常查询操作日志_验证条件搜索",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /system/operation-log\n7. [等待] 1000\n8. [选择] .el-select[placeholder*=\"操作类型\"]::LOGIN\n9. [点击] 搜索\n10. [等待] 1000\n11. [断言] LOGIN 可见\n12. [DB] SELECT operation_type,COUNT(*) as cnt FROM tbl_operation_log WHERE operation_type='LOGIN' GROUP BY operation_type",
    "1-5. 登录成功\n6-7. 导航到操作日志页面\n8-9. 按操作类型LOGIN搜索\n10. 页面刷新\n11. 列表显示LOGIN类型的日志\n12. 数据库返回LOGIN类型的日志数量",
    "P3"))

test_cases.append((oplog_module, oplog_req,
    "正常查看日志详情_验证JSON数据展示",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /system/operation-log\n7. [等待] 1000\n8. [点击] 第一条日志行的查看按钮\n9. [等待] 500\n10. [断言] before_json 可见 或 after_json 可见 或 变更前 可见\n11. [截图] 日志详情JSON展示",
    "1-5. 登录成功\n6-7. 导航到操作日志页面\n8. 详情对话框打开\n9-10. 显示变更前/变更后JSON数据\n11. 截图保存成功",
    "P3"))

test_cases.append((oplog_module, oplog_req,
    "正常操作日志_验证新增入居操作自动记录",
    AUTH_PRECOND + "\n4. 存在空房间和可用员工",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /resident/records\n7. [点击] 入居登録\n8. [选择] 员工和房间并提交\n9. [等待] 1000\n10. [导航] /system/operation-log\n11. [等待] 1000\n12. [点击] 搜索\n13. [断言] CHECKIN 可见\n14. [DB] SELECT * FROM tbl_operation_log WHERE operation_type='CHECKIN' ORDER BY create_time DESC LIMIT 1",
    "1-5. 登录成功\n6-8. 执行入居操作\n9. 入居成功\n10-12. 导航到操作日志查看\n13. 日志列表中显示CHECKIN类型记录\n14. 数据库返回最新入居操作日志",
    "P2"))

test_cases.append((oplog_module, oplog_req,
    "正常操作日志_验证失败操作记录日志",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::wrongpassword\n4. [点击] 登录\n5. [等待] 1000\n6. [导航] /system/operation-log\n7. [等待] 1000\n8. [选择] .el-select[placeholder*=\"结果\"]::失败\n9. [点击] 搜索\n10. [等待] 1000\n11. [断言] LOGIN 可见\n12. [DB] SELECT * FROM tbl_operation_log WHERE result_status=0 ORDER BY create_time DESC LIMIT 1",
    "1-4. 执行错误密码登录\n5. 登录失败\n6-7. 导航到操作日志页面\n8-9. 按失败结果搜索\n10. 页面刷新\n11. 显示失败的登录日志\n12. 数据库返回失败操作日志记录",
    "P2"))

# =====================================================================
# 12. 変更履歴 (変更履歴)
# =====================================================================
chglog_module = "変更履歴"
chglog_req = "15. ログおよび監査設計 / 8. 変更履歴API（参照専用）"

test_cases.append((chglog_module, chglog_req,
    "正常查询变更履历列表_验证默认加载",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /system/change-history\n7. [等待] 1000\n8. [断言] 变更履历 可见\n9. [DB] SELECT COUNT(*) as cnt FROM tbl_change_history",
    "1-5. 登录成功\n6-7. 导航到变更履历页面\n8. 页面标题显示变更履历\n9. 数据库返回变更履历总数",
    "P3"))

test_cases.append((chglog_module, chglog_req,
    "正常查询变更履历_验证按表名搜索",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /system/change-history\n7. [等待] 1000\n8. [输入] input[placeholder*=\"表名\"]::tbl_employee\n9. [点击] 搜索\n10. [等待] 1000\n11. [断言] tbl_employee 可见\n12. [DB] SELECT table_name,COUNT(*) as cnt FROM tbl_change_history WHERE table_name='tbl_employee' GROUP BY table_name",
    "1-5. 登录成功\n6-7. 导航到变更履历页面\n8-9. 按表名tbl_employee搜索\n10. 页面刷新\n11. 列表显示该表的变更履历\n12. 数据库返回该表的变更记录数",
    "P3"))

test_cases.append((chglog_module, chglog_req,
    "正常查看变更履历详情_验证前后数据对比",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /system/change-history\n7. [等待] 1000\n8. [点击] 第一条履历行的查看按钮\n9. [等待] 500\n10. [断言] 旧数据 可见 或 新数据 可见 或 old 可见\n11. [截图] 变更前后对比",
    "1-5. 登录成功\n6-7. 导航到变更履历页面\n8. 详情对话框打开\n9-10. 显示变更前后数据对比\n11. 截图保存成功",
    "P3"))

test_cases.append((chglog_module, chglog_req,
    "正常变更履历_验证修改数据后自动生成履历",
    AUTH_PRECOND + "\n4. 存在可修改的部门记录",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/department\n7. [等待] 1000\n8. [点击] 目标部门行编辑按钮\n9. [修改] 部门名称\n10. [点击] 确定\n11. [等待] 1000\n12. [导航] /system/change-history\n13. [等待] 1000\n14. [输入] input[placeholder*=\"表名\"]::tbl_department\n15. [点击] 搜索\n16. [等待] 1000\n17. [断言] UPDATE 可见\n18. [DB] SELECT * FROM tbl_change_history WHERE table_name='tbl_department' AND operation_type='UPDATE' ORDER BY create_time DESC LIMIT 1",
    "1-5. 登录成功\n6-10. 执行部门更新操作\n11. 更新成功\n12-16. 导航到变更履历查询\n17. 显示UPDATE类型的变更记录\n18. 数据库返回最新变更记录",
    "P2"))

# =====================================================================
# 13. 備品管理 (備品管理)
# =====================================================================
fixture_module = "備品マスタ管理"
fixture_req = "7. 備品管理 / 7.1 備品マスタ"

test_cases.append((fixture_module, fixture_req,
    "正常查询备品列表_验证默认加载",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/fixture\n7. [等待] 1000\n8. [断言] 备品管理 可见\n9. [DB] SELECT COUNT(*) as cnt FROM tbl_fixture",
    "1-5. 登录成功\n6-7. 导航到备品管理页面\n8. 页面标题显示备品管理\n9. 数据库返回备品总数",
    "P3"))

test_cases.append((fixture_module, fixture_req,
    "正常创建备品_验证新增备品成功",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/fixture\n7. [等待] 1000\n8. [点击] 新增\n9. [等待] 500\n10. [输入] .el-dialog input[placeholder*=\"备品名称\"]::测试备品\n11. [选择] .el-dialog .el-select[placeholder*=\"备品种别\"]::家具\n12. [点击] 确定\n13. [等待] 1000\n14. [断言] 测试备品 可见\n15. [DB] SELECT name,type FROM tbl_fixture WHERE name='测试备品'",
    "1-5. 登录成功\n6-7. 导航到备品管理页面\n8. 新增对话框打开\n9-11. 填写备品信息\n12. 提交创建\n13. 处理请求\n14. 列表显示新备品\n15. 数据库返回新备品记录",
    "P3"))

test_cases.append((fixture_module, fixture_req,
    "正常更新备品_验证更新成功",
    AUTH_PRECOND + "\n4. 已存在测试备品",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/fixture\n7. [等待] 1000\n8. [点击] 目标备品行编辑按钮\n9. [等待] 500\n10. [输入] .el-dialog input[placeholder*=\"备品名称\"]::更新后的备品名\n11. [点击] 确定\n12. [等待] 1000\n13. [断言] 更新后的备品名 可见\n14. [DB] SELECT name FROM tbl_fixture WHERE id={id}",
    "1-5. 登录成功\n6-7. 导航到备品管理页面\n8. 编辑对话框打开\n9-10. 修改备品名称\n11. 提交更新\n12. 处理请求\n13. 列表显示更新后的名称\n14. 数据库返回更新后的备品名",
    "P3"))

test_cases.append((fixture_module, fixture_req,
    "正常删除备品_验证删除成功",
    AUTH_PRECOND + "\n4. 已存在测试备品",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/fixture\n7. [等待] 1000\n8. [点击] 目标备品行删除按钮\n9. [等待] 500\n10. [点击] 确认\n11. [等待] 1000\n12. [断言] 删除成功 可见\n13. [DB] SELECT COUNT(*) as cnt FROM tbl_fixture WHERE id={id}",
    "1-5. 登录成功\n6-7. 导航到备品管理页面\n8. 删除确认对话框打开\n9-10. 确认删除\n11. 处理请求\n12. 页面提示删除成功\n13. 数据库返回0条记录",
    "P3"))

# =====================================================================
# 14. Excelインポート (Excelインポート)
# =====================================================================
import_module = "Excelインポート"
import_req = "9. Excelデータ取り込み / 10. Excelインポートおよびバッチ設計"

test_cases.append((import_module, import_req,
    "正常Excelインポート_验证寮マスタ导入成功",
    AUTH_PRECOND + "\n4. 准备测试Excel文件（寮マスタ.xlsx，包含有效的寮数据）",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /import\n7. [等待] 1000\n8. [选择] .el-select[placeholder*=\"导入种别\"]::寮マスタ\n9. [上传] input[type=\"file\"]::D:\\test\\寮マスタ.xlsx\n10. [等待] 3000\n11. [断言] 导入成功 可见\n12. [DB] SELECT import_status,success_count,total_count FROM tbl_import_log ORDER BY create_time DESC LIMIT 1",
    "1-5. 登录成功\n6-7. 导航到Excelインポート页面\n8. 选择导入种别\n9. 上传Excel文件\n10. 执行导入处理\n11. 页面提示导入成功\n12. 数据库返回导入日志，status=1(成功)",
    "P2"))

test_cases.append((import_module, import_req,
    "异常Excelインポート_验证格式错误提示",
    AUTH_PRECOND + "\n4. 准备格式错误的Excel文件",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /import\n7. [等待] 1000\n8. [选择] .el-select[placeholder*=\"导入种别\"]::寮マスタ\n9. [上传] input[type=\"file\"]::D:\\test\\不正フォーマット.xlsx\n10. [等待] 3000\n11. [断言] エラー 可见 或 失败 可见\n12. [DB] SELECT import_status,error_details FROM tbl_import_log ORDER BY create_time DESC LIMIT 1",
    "1-5. 登录成功\n6-7. 导航到Excelインポート页面\n8-9. 上传格式错误的文件\n10. 执行导入处理\n11. 页面提示错误信息\n12. 数据库返回导入日志，status=2(部分失败)或3(失败)",
    "P2"))

test_cases.append((import_module, import_req,
    "异常Excelインポート_验证重复数据校验",
    AUTH_PRECOND + "\n4. 准备包含已存在数据的Excel文件",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /import\n7. [等待] 1000\n8. [选择] .el-select[placeholder*=\"导入种别\"]::部屋マスタ\n9. [上传] input[type=\"file\"]::D:\\test\\重复数据.xlsx\n10. [等待] 3000\n11. [断言] 重复 可见 或 已存在 可见\n12. [DB] SELECT error_details FROM tbl_import_log ORDER BY create_time DESC LIMIT 1",
    "1-5. 登录成功\n6-7. 导航到Excelインポート页面\n8-9. 上传含重复数据的文件\n10. 执行导入处理\n11. 页面提示重复数据错误\n12. 数据库返回导入日志含错误详情",
    "P2"))

test_cases.append((import_module, import_req,
    "正常Excelインポート_验证导入履歴一覧",
    AUTH_PRECOND + "\n4. 已有至少一次导入记录",
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /import\n7. [等待] 1000\n8. [断言] 履歴 可见\n9. [点击] 导入履歴选项卡\n10. [等待] 1000\n11. [断言] 成功 可见 或 失败 可见\n12. [DB] SELECT file_name,import_type,import_status FROM tbl_import_log ORDER BY create_time DESC",
    "1-5. 登录成功\n6-7. 导航到Excelインポート页面\n8. 页面显示履歴选项卡\n9-10. 查看导入履歴\n11. 履歴列表显示导入结果\n12. 数据库返回最新导入记录",
    "P3"))

test_cases.append((import_module, import_req,
    "异常Excelインポート_验证未选择文件提示",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /import\n7. [等待] 1000\n8. [点击] 导入按钮\n9. [等待] 1000\n10. [断言] 选择文件 可见 或 ファイル 可见",
    "1-5. 登录成功\n6-7. 导航到Excelインポート页面\n8. 未选择文件直接点击导入\n9. 系统校验触发\n10. 提示请选择文件",
    "P3"))

# =====================================================================
# 15. 帳票・統計 (レポート)
# =====================================================================
report_module = "帳票・統計"
report_req = "5.3 寮費プロセス / 帳票・統計画面"

test_cases.append((report_module, report_req,
    "正常レポート表示_验证ダッシュボードサマリー",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /report\n7. [等待] 1000\n8. [断言] 総寮数 可见 或 サマリー 可见\n9. [DB] SELECT COUNT(*) as total_dorm FROM tbl_dormitory WHERE status=1",
    "1-5. 登录成功\n6-7. 导航到レポート页面\n8. 页面显示サマリーカード\n9. 数据库返回运营中寮总数",
    "P3"))

test_cases.append((report_module, report_req,
    "正常レポート表示_验证寮別稼働率",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /report\n7. [等待] 1000\n8. [点击] 寮別稼働率 选项卡\n9. [等待] 1000\n10. [断言] 稼働率 可见 或 占有率 可见\n11. [截图] 寮別稼働率レポート",
    "1-5. 登录成功\n6-7. 导航到レポート页面\n8-9. 切换到寮別稼働率选项卡\n10. 页面显示稼働率表格及进度条\n11. 截图保存成功",
    "P3"))

test_cases.append((report_module, report_req,
    "正常レポート表示_验证請求サマリー",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /report\n7. [等待] 1000\n8. [点击] 請求サマリー 选项卡\n9. [等待] 1000\n10. [断言] 請求 可见 或 金额 可见\n11. [截图] 請求サマリーレポート",
    "1-5. 登录成功\n6-7. 导航到レポート页面\n8-9. 切换到請求サマリー选项卡\n10. 页面显示請求汇总数据\n11. 截图保存成功",
    "P3"))

# =====================================================================
# 16. ホーム画面 (ホーム)
# =====================================================================
home_module = "ホーム画面"
home_req = "11. 画面およびインタラクション設計 / ホーム画面"

test_cases.append((home_module, home_req,
    "正常ホーム表示_验证仪表盘加载",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [断言URL] /home\n7. [断言] 欢迎 可见\n8. [断言] 快捷链接 可见 或 クイックリンク 可见\n9. [截图] ホーム仪表盘",
    "1-4. 登录成功\n5-6. 页面跳转到/home\n7. 页面显示欢迎信息\n8. 页面显示功能快捷链接卡片\n9. 截图保存成功",
    "P1"))

test_cases.append((home_module, home_req,
    "正常ホーム表示_验证快捷链接导航",
    AUTH_PRECOND,
    "功能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [断言URL] /home\n7. [点击] 入居履历 快捷链接\n8. [等待] 1000\n9. [断言URL] /resident/records\n10. [断言] 入居履历 可见",
    "1-5. 登录成功\n6. URL为/home\n7. 点击入居履历快捷链接\n8-9. 页面跳转到入居履历页面\n10. 页面标题显示入居履历",
    "P2"))

# =====================================================================
# 17. ホーム画面 - パフォーマンステスト
# =====================================================================
perf_module = "非機能要件"
perf_req = "18. 非機能要件 / 13. パフォーマンス最適化設計"

test_cases.append((perf_module, perf_req,
    "性能测试_验证API响应时间在1秒以内",
    AUTH_PRECOND + "\n4. 使用性能测试工具（如Postman或curl）",
    "性能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [API] GET /api/dormitories/list (JWT認証)\n7. [API] GET /api/resident-records/list (JWT認証)\n8. [API] GET /api/employees/list (JWT認証)\n9. [断言] 各API响应时间 < 1000ms",
    "1-5. 登录成功获取JWT Token\n6. 寮一覧API响应时间<1000ms\n7. 入居履歴一覧API响应时间<1000ms\n8. 社員一覧API响应时间<1000ms\n9. 所有API符合性能指标",
    "P4"))

test_cases.append((perf_module, perf_req,
    "性能测试_验证画面响应时间在3秒以内",
    AUTH_PRECOND,
    "性能测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [导航] /resident/records\n6. [等待] 5000\n7. [断言] 页面加载时间 < 3000ms\n8. [截图] 页面加载计时",
    "1-3. 登录\n4. 登录成功\n5. 导航到入居履歴页面\n6. 等待页面加载\n7. 页面在3秒内加载完成\n8. 截图保存",
    "P4"))

test_cases.append((perf_module, perf_req,
    "并发测试_验证多用户同时操作不冲突",
    AUTH_PRECOND + "\n4. 准备多个测试用户账号\n5. 准备多个可用房间",
    "性能测试",
    "1. [同时操作] 用户A入居登録（房间A）\n2. [同时操作] 用户B入居登録（房间B）\n3. [等待] 3000\n4. [DB] SELECT is_active,employee_id,room_id FROM tbl_resident_record WHERE room_id IN (A,B) AND is_active=1\n5. [断言] 两个入居记录均成功\n6. [DB] SELECT current_occupancy FROM tbl_room WHERE id IN (A,B)",
    "1-2. 两个用户同时入居操作\n3. 处理完成\n4. 数据库两条入居记录均is_active=1\n5. 互不冲突\n6. 两个房间入居人数均正确更新",
    "P4"))

# =====================================================================
# 18. セキュリティテスト
# =====================================================================
sec_module = "非機能要件"
sec_req = "16. セキュリティ設計"

test_cases.append((sec_module, sec_req,
    "安全测试_验证未认证访问重定向到登录页",
    AUTH_PRECOND,
    "安全测试",
    "1. [导航] /system/user（未登录状态）\n2. [等待] 1000\n3. [断言URL] /login\n4. [导航] /api/dormitories/list（未携带Token）\n5. [断言] HTTP 401 可见 或 未认证 可见",
    "1-2. 直接访问需认证页面\n3. 自动重定向到登录页\n4. 直接调用API不带Token\n5. 返回401未认证错误",
    "P1"))

test_cases.append((sec_module, sec_req,
    "安全测试_验证SQL注入防护",
    AUTH_PRECOND,
    "安全测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin' OR '1'='1\n3. [输入] input[placeholder*=\"密码\"]::' OR '1'='1\n4. [点击] 登录\n5. [等待] 1000\n6. [断言] 用户不存在 可见 或 密码错误 可见\n7. [断言URL] /login",
    "1. 登录页面加载\n2-3. 输入SQL注入语句\n4. 点击登录\n5. 处理请求\n6. SQL注入失败，显示用户不存在或密码错误\n7. URL仍为/login（未跳转）",
    "P1"))

test_cases.append((sec_module, sec_req,
    "安全测试_验证XSS脚本过滤",
    AUTH_PRECOND,
    "安全测试",
    "1. [导航] /login\n2. [输入] input[placeholder*=\"用户\"]::admin\n3. [输入] input[placeholder*=\"密码\"]::admin123\n4. [点击] 登录\n5. [等待] 2000\n6. [导航] /master/department\n7. [点击] 新增\n8. [输入] .el-dialog input[placeholder*=\"部门名称\"]::<script>alert('XSS')</script>\n9. [点击] 确定\n10. [等待] 1000\n11. [断言] <script> 不可见 或 已转义 可见\n12. [DB] SELECT dept_name FROM tbl_department WHERE dept_name LIKE '%XSS%'",
    "1-5. 登录成功\n6-7. 导航到部门管理新增\n8. 输入XSS脚本\n9. 提交\n10. 处理请求\n11. 页面显示转义后的内容（不是可执行的脚本）\n12. 数据库存储原始数据或转义后数据",
    "P1"))


# ============================================================
# データ準備完了
# ============================================================

def generate_files():
    """ExcelとMarkdownファイルを生成"""
    import datetime
    
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Excelファイル生成
    generate_excel(now)
    
    # Markdownファイル生成
    generate_markdown(now)


def generate_excel(now):
    """Excelファイルを生成"""
    wb = Workbook()
    ws = wb.active
    ws.title = "テストケース"
    
    # スタイル定義
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Microsoft YaHei", size=10)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 列幅設定
    col_widths = [18, 25, 40, 35, 14, 60, 50, 8]
    headers = ["所属模块", "相关研发需求", "用例名称", "前置条件", "用例类型", "测试步骤", "预期结果", "优先级"]
    
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # データ行
    for row_idx, tc in enumerate(test_cases, 2):
        for col_idx, value in enumerate(tc, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # フィルター設定
    ws.auto_filter.ref = f"A1:H{len(test_cases)+1}"
    
    # フリーズペイン（ヘッダー行固定）
    ws.freeze_panes = "A2"
    
    # 行の高さ
    ws.row_dimensions[1].height = 30
    
    filepath = f"{OUTPUT_DIR}\\{PROJECT_NAME}_{RANDOM_SUFFIX}.xlsx"
    wb.save(filepath)
    print(f"Excel saved: {filepath}")


def generate_markdown(now):
    """Markdownファイルを生成"""
    # モジュールごとにグループ化
    module_groups = {}
    for tc in test_cases:
        module = tc[0]
        if module not in module_groups:
            module_groups[module] = []
        module_groups[module].append(tc)
    
    # 用例汇总统计
    module_summary = {}
    for module, cases in module_groups.items():
        p1 = sum(1 for c in cases if c[7] == "P1")
        p2 = sum(1 for c in cases if c[7] == "P2")
        p3 = sum(1 for c in cases if c[7] == "P3")
        p4 = sum(1 for c in cases if c[7] == "P4")
        module_summary[module] = (len(cases), p1, p2, p3, p4)
    
    total_cases = len(test_cases)
    total_p1 = sum(v[1] for v in module_summary.values())
    total_p2 = sum(v[2] for v in module_summary.values())
    total_p3 = sum(v[3] for v in module_summary.values())
    total_p4 = sum(v[4] for v in module_summary.values())
    
    lines = []
    lines.append(f"# {PROJECT_NAME} 测试用例\n")
    lines.append("> 生成时间: " + now)
    lines.append(f"> 需求来源: {PRD_DOC_NAME} v1.0")
    lines.append(f"> 测试环境: {FRONTEND_URL} + {BACKEND_URL}")
    lines.append(f"> 数据库: {DB_INFO}")
    lines.append(f"> 对应Excel: {PROJECT_NAME}_{RANDOM_SUFFIX}.xlsx\n")
    lines.append("---\n")
    
    # 用例汇总表
    lines.append("## 用例汇总\n")
    lines.append("| 模块 | 用例数 | P1 | P2 | P3 | P4 |")
    lines.append("|------|--------|----|----|----|----|")
    for module in module_groups:
        cnt, p1, p2, p3, p4 = module_summary[module]
        lines.append(f"| {module} | {cnt} | {p1} | {p2} | {p3} | {p4} |")
    lines.append(f"| **合计** | **{total_cases}** | **{total_p1}** | **{total_p2}** | **{total_p3}** | **{total_p4}** |\n")
    lines.append("---\n")
    
    # 详细用例
    lines.append("## 详细用例\n")
    
    tc_num = 1
    for module_idx, (module, cases) in enumerate(module_groups.items(), 1):
        lines.append(f"### {module_idx}. {module}\n")
        
        for case_idx, tc in enumerate(cases, 1):
            lines.append(f"#### TC-{tc_num:02d}: {tc[2]}\n")
            lines.append("| 属性 | 内容 |")
            lines.append("|------|------|")
            lines.append(f"| **所属模块** | {tc[0]} |")
            lines.append(f"| **相关需求** | {tc[1]} |")
            lines.append(f"| **优先级** | {tc[7]} |")
            lines.append(f"| **用例类型** | {tc[4]} |")
            lines.append(f"| **前置条件** | {tc[3].replace(chr(10), '<br>')} |\n")
            
            lines.append("**测试步骤与预期结果**:\n")
            lines.append("| 步骤 | 操作 | 预期结果 |")
            lines.append("|------|------|----------|")
            
            steps = tc[5].split('\n')
            expected = tc[6].split('\n')
            max_len = max(len(steps), len(expected))
            
            for i in range(max_len):
                step_text = steps[i] if i < len(steps) else ""
                expected_text = expected[i] if i < len(expected) else ""
                lines.append(f"| {i+1} | {step_text} | {expected_text} |")
            
            lines.append("")
            lines.append("---\n")
            tc_num += 1
    
    filepath = f"{OUTPUT_DIR}\\{PROJECT_NAME}_{RANDOM_SUFFIX}.md"
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f"Markdown saved: {filepath}")


if __name__ == "__main__":
    generate_files()
    print(f"\nTotal test cases generated: {len(test_cases)}")
    print(f"File prefix: {PROJECT_NAME}_{RANDOM_SUFFIX}")
